import io
import os
import shutil
import tempfile

import openpyxl
from weasyprint import CSS, HTML
from xlsx2html import xlsx2html

from src.settings.models import ReceiptTemplate


def _resolve_template(template_id=None):
    if template_id and str(template_id).isdigit():
        t = ReceiptTemplate.objects.filter(pk=int(template_id)).first()
        if t and t.file:
            return t
    return (
        ReceiptTemplate.objects.filter(is_default=True).first()
        or ReceiptTemplate.objects.order_by("created_at").first()
    )


def generate_receipt_xlsx(receipt, *, template_id=None) -> bytes | None:
    template = _resolve_template(template_id)
    if not template or not template.file:
        return None
    data = _collect(receipt)
    return _fill_template(data, template)


def generate_receipt_pdf(receipt, *, template_id=None) -> bytes | None:
    xlsx_bytes = generate_receipt_xlsx(receipt, template_id=template_id)
    if not xlsx_bytes:
        return None

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name

    try:
        html_buf = io.StringIO()
        xlsx2html(tmp_path, html_buf)
        html_content = html_buf.getvalue()
        return HTML(string=html_content).write_pdf(
            stylesheets=[
                CSS(string="@page { size: A4 landscape; margin: 1cm; }")
            ]
        )
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def _collect(receipt) -> dict:
    lines = list(
        receipt.lines.select_related("service", "service__unit_of_measurement")
    )
    flat = receipt.flat if receipt.flat_id else None
    owner = flat.owner if flat and flat.owner_id else None
    bank_book = receipt.bank_book if receipt.bank_book_id else None

    balance = 0.0
    if bank_book:
        from src.finances.views.bank_books import bank_book_balance

        balance = bank_book_balance(bank_book)

    period = ""
    if receipt.period_from and receipt.period_to:
        period = (
            f"{receipt.period_from.strftime('%d.%m.%Y')}"
            f" - {receipt.period_to.strftime('%d.%m.%Y')}"
        )

    service_rows = []
    for ln in lines:
        svc = ln.service
        unit = (
            svc.unit_of_measurement.title
            if svc and svc.unit_of_measurement_id
            else ""
        )
        service_rows.append(
            {
                "service": svc.title if svc else "",
                "tariff": receipt.tariff.title if receipt.tariff_id else "",
                "unit": unit,
                "amount": ln.amount,
                "price": ln.price,
                "total": round(ln.price * ln.amount, 2),
            }
        )

    return {
        "completed": (
            "Zaksięgowano" if receipt.completed else "Nie zaksięgowano"
        ),
        "status": receipt.get_status_display(),
        "period": period,
        "owner": owner.get_full_name() if owner else "",
        "phone": (
            str(owner.phone_number)
            if owner and getattr(owner, "phone_number", None)
            else ""
        ),
        "bank_book_number": bank_book.random_number if bank_book else "",
        "house": flat.house.title if flat and flat.house_id else "",
        "flat_number": str(flat.number) if flat else "",
        "section": flat.section.title if flat and flat.section_id else "",
        "tariff": receipt.tariff.title if receipt.tariff_id else "",
        "balance": round(balance, 2),
        "service_rows": service_rows,
        "total": round(sum(r["total"] for r in service_rows), 2),
    }


def _fill_template(data: dict, template) -> bytes:
    """Fill the template adaptively by scanning for label cells and column headers.

    No cell addresses are hardcoded — the code reads the template structure and
    adapts to any layout, so renaming rows or moving columns never breaks it.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        shutil.copy(template.file.path, tmp_path)
        wb = openpyxl.load_workbook(tmp_path, keep_vba=False)
        ws = wb.active

        # ── 1. Fill header info: find each label, write value one cell to the right ──
        # Normalised label text (stripped, colon removed, lower) → data key
        LABEL_MAP = {
            "проведена": "completed",
            "статус": "status",
            "период": "period",
            "владелец": "owner",
            "лицевой счет": "bank_book_number",
            "телефон": "phone",
            "дом": "house",
            "квартира": "flat_number",
            "секция": "section",
            "тариф": "tariff",
        }
        # Collect first, write second — avoids mutating cells that the iterator
        # will visit next (which would cascade the value into adjacent cells).
        label_writes: list[tuple[int, int, object]] = []
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                key = cell.value.strip().rstrip(":").lower()
                if key in LABEL_MAP:
                    label_writes.append(
                        (cell.row, cell.column + 1, data[LABEL_MAP[key]])
                    )
        for r, c, val in label_writes:
            ws.cell(row=r, column=c, value=val)

        # ── 2. Locate the service table header row and map column roles ──
        # Each keyword that, if found anywhere inside the cell text, identifies the column role.
        # Listed most-specific first to avoid "цена за ед." matching both "цена" and "ед."
        COL_KEYWORDS = [
            ("количество", "amount"),
            ("расход", "amount"),
            ("ед.изм", "unit"),
            ("ед. изм", "unit"),
            ("единиц", "unit"),
            ("стоимость", "cost"),
            ("сумма", "cost"),
            ("цена", "price"),
            ("услуга", "service"),
            ("наименован", "service"),
            ("#", "idx"),
        ]

        header_row_num = None
        col_map: dict[str, int] = {}  # role → 1-based column index

        for row in ws.iter_rows():
            found: dict[str, int] = {}
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                val_lower = cell.value.strip().lower()
                for keyword, role in COL_KEYWORDS:
                    if keyword in val_lower and role not in found:
                        found[role] = cell.column
                        break  # each cell gets one role
            # A valid service-table header must contain at least "service" + "amount"
            if "service" in found and "amount" in found:
                header_row_num = row[0].row
                col_map = found
                break

        # ── 3. Fill service rows (insert extra rows as needed) ────────────────────
        if header_row_num and col_map:
            start = header_row_num + 1
            for i, svc in enumerate(data["service_rows"]):
                r = start + i
                if i > 0:
                    ws.insert_rows(r)  # shifts everything below down by 1
                if "idx" in col_map:
                    ws.cell(row=r, column=col_map["idx"], value=i + 1)
                if "service" in col_map:
                    ws.cell(
                        row=r, column=col_map["service"], value=svc["service"]
                    )
                if "amount" in col_map:
                    ws.cell(
                        row=r, column=col_map["amount"], value=svc["amount"]
                    )
                if "unit" in col_map:
                    ws.cell(row=r, column=col_map["unit"], value=svc["unit"])
                if "price" in col_map:
                    ws.cell(row=r, column=col_map["price"], value=svc["price"])
                if "cost" in col_map:
                    ws.cell(row=r, column=col_map["cost"], value=svc["total"])

        # ── 4. Fill total — find "Итого" after row insertions (row may have shifted) ──
        # Collect position first, write separately.
        itogo_pos: tuple[int, int] | None = None
        for row in ws.iter_rows():
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and "итого" in cell.value.strip().lower()
                ):
                    itogo_pos = (cell.row, cell.column + 1)
                    break
            if itogo_pos:
                break
        if itogo_pos:
            ws.cell(row=itogo_pos[0], column=itogo_pos[1], value=data["total"])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
